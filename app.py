from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

orden_prioridad = ['Marca', 'Resistencia', 'Seguridad', 'Pase', 'Gambeta', 'Velocidad', 'Cabeza', 'Potencia', 'Definicion']
archivo_guardado = "jugadores_guardados.xlsx"

tabla_jugadores = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/guardar_jugador', methods=['POST'])
def guardar_jugador():
    nombre = request.form['nombre']
    habilidades = [int(request.form[h]) for h in orden_prioridad]
    
    try:
        wb = load_workbook(archivo_guardado)
        ws = wb.active
        if nombre in tabla_jugadores:
            idx = list(tabla_jugadores.keys()).index(nombre) + 2
            for col, valor in zip(range(2, len(orden_prioridad) + 2), habilidades):
                ws.cell(row=idx, column=col, value=valor)
            tabla_jugadores[nombre] = habilidades
        else:
            ws.append([nombre] + habilidades)
            tabla_jugadores[nombre] = habilidades
        
        wb.save(archivo_guardado)
        return jsonify({'mensaje': 'Jugador guardado correctamente'})
    except Exception as e:
        return jsonify({'error': str(e)})

# Agrega más rutas y funciones según sea necesario

if __name__ == '__main__':
    app.run(debug=True)
